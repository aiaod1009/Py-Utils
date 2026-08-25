"""
麦谷上游 vs STS中转 GLM 模型对比测试脚本
=========================================
对比两套端点（麦谷直连 vs STS中转）在 glm-5.1 / glm-5.2 两个模型上
于 4K/16K/64K/128K 上下文下的性能差异：
  - TTFT (Time to First Token) 首 token 延迟
  - TPOT (Time per Output Token) 每 token 输出时间
  - TPS  (Tokens Per Second)   每秒输出 token 数
  - 任务完成质量评分
  - 成功率
输出对比 Excel 报告。

用法:
  python maigu_sts_compare.py          # 完整测试
  python maigu_sts_compare.py --quick  # 快速模式（每组合跑1次）
  python maigu_sts_compare.py --resume  # 断点续跑
"""

from __future__ import annotations

import json
import os
import statistics
import time
import traceback
from collections import defaultdict
from dataclasses import dataclass, field, asdict
from datetime import datetime
from typing import Optional

import requests

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[警告] openpyxl 未安装，无法生成 Excel。请执行: pip install openpyxl")


# 全局 Session：禁用系统代理读取，避免 Clash 等代理拦截导致 SSL EOF
API_SESSION = requests.Session()
API_SESSION.trust_env = False


# ============================================================
# .env 加载
# ============================================================

def _load_env() -> dict[str, str]:
    env: dict[str, str] = {}
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    if os.path.isfile(env_path):
        with open(env_path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip().strip('"').strip("'")
    return env


ENV = _load_env()

MAIGU_BASE_URL = ENV.get("MAIGU_BASE_URL", "https://ka.ent-pilot.com")
MAIGU_API_KEY = ENV.get("MAIGU_API_KEY", "")
STS_BASE_URL = ENV.get("STS_BASE_URL", "https://www.sts-token.com/v1")
STS_API_KEY = ENV.get("STS_API_KEY", "")

if not MAIGU_API_KEY or not STS_API_KEY:
    print("[错误] 请在 .env 中设置 MAIGU_API_KEY 和 STS_API_KEY")
    exit(1)


# ============================================================
# 配置
# ============================================================

# 端点配置（base_url 兼容带或不带 /v1 后缀）
ENDPOINTS: list[dict] = [
    {"name": "麦谷上游", "base_url": MAIGU_BASE_URL, "api_key": MAIGU_API_KEY},
    {"name": "STS中转", "base_url": STS_BASE_URL, "api_key": STS_API_KEY},
]

# 测试模型
MODELS: list[str] = ["glm-5.1", "glm-5.2"]

# 上下文测试条件
CONDITIONS: list[dict] = [
    {"name": "4K",   "input_tokens": 3072,   "max_tokens": 1024, "samples": 3},
    {"name": "16K",  "input_tokens": 14000,  "max_tokens": 2048, "samples": 3},
    {"name": "64K",  "input_tokens": 62000,  "max_tokens": 2048, "samples": 2},
    {"name": "128K", "input_tokens": 126000, "max_tokens": 2048, "samples": 2},
]

# 单次任务评估问题（每样本只问 1 题，专注于 TTFT/TPOT）
# 注：GLM 在长 prompt 下会强制开启 thinking，问题过复杂会导致 4K(max_tokens=1024) 思考占满
# 因此使用单一聚焦任务，保证 4K 也有空间输出可见内容
EVALUATION_QUESTION = (
    "请基于上述文章，用不超过150字概括文章的核心观点，并简要说明 Transformer 架构的意义。"
)

REQUEST_TIMEOUT = 300  # 秒

# 背景语料（1 token ≈ 1.4 中文字符）
BASE_TEXT = (
    "人工智能（Artificial Intelligence，简称AI）是计算机科学的一个重要分支，"
    "它致力于研究、开发用于模拟、延伸和扩展人类智能的理论、方法、技术及应用系统。"
    "自1956年达特茅斯会议正式提出人工智能概念以来，该领域经历了多次起伏，"
    "从早期的符号主义到专家系统，再到统计机器学习和深度学习的崛起。"
    "近年来，随着深度学习技术的突破、大数据基础设施的发展以及GPU/TPU等计算能力的提升，"
    "人工智能在图像识别、自然语言处理、语音识别、自动驾驶、医疗诊断等领域取得了显著进展。"
    "大型语言模型（LLM）如GPT系列、GLM系列、Claude系列等的出现，更是将人工智能推向了一个新的高度，"
    "这些模型通过对海量文本数据的学习，展现出了强大的语言理解和生成能力。"
    "在自然语言处理领域，预训练-微调范式已成为主流，BERT、GPT、T5等模型架构层出不穷。"
    "Transformer架构的自注意力机制使得模型能够捕捉长距离依赖关系，这是传统RNN/LSTM难以做到的。"
    "强化学习与人类反馈（RLHF）技术的引入，使得大语言模型能够更好地对齐人类意图和价值观。"
    "然而，人工智能的发展也面临着诸多挑战，包括数据隐私保护、算法公平性与偏见、能源消耗、"
    "以及如何确保AI系统的安全性、可控性和可解释性等问题。"
    "在未来，人工智能将继续深刻影响各行各业，从医疗健康到教育培训，"
    "从金融科技到智能制造，从智慧城市到环境保护，AI的应用前景广阔无垠。"
    "同时，我们也需要关注AI伦理和治理，确保技术的发展真正造福人类社会。"
    "量子计算与人工智能的结合可能带来革命性突破，量子机器学习有望解决经典计算难以处理的问题。"
    "边缘计算与联邦学习则让AI能力延伸到更多终端设备，同时保护用户数据隐私。"
    "多模态学习使得AI能够同时处理文本、图像、音频、视频等多种信息形式，"
    "这为实现更接近人类的感知和理解能力奠定了基础。"
    "AI Agent技术的发展使得大模型能够自主规划、调用工具、执行多步骤任务，"
    "这标志着人工智能从被动应答向主动执行的重要转变。"
    "在实际应用中，AI辅助编程工具如GitHub Copilot、Cursor等极大地提高了开发者的生产效率。"
    "AI在科学研究中的应用也日益广泛，从蛋白质结构预测到新材料发现，从气候模拟到药物研发。"
    "计算机视觉领域，目标检测、语义分割、图像生成（如Stable Diffusion、DALL-E）等技术日趋成熟。"
    "语音技术方面，语音识别、语音合成、声纹识别等已广泛应用于智能音箱、客服系统等场景。"
    "推荐系统借助深度学习，在电商、短视频、社交媒体等平台实现了精准的个性化推荐。"
    "知识图谱与大型语言模型的结合，有望解决大模型的幻觉问题和知识更新问题。"
    "检索增强生成（RAG）技术通过引入外部知识库，显著提升了模型回答的准确性和时效性。"
    "提示工程（Prompt Engineering）和上下文学习（In-Context Learning）成为高效使用大模型的关键技能。"
    "模型压缩技术如量化、剪枝、蒸馏等，使得大模型能够在资源受限的设备上高效运行。"
)

REPORT_FILE = "maigu_sts_report.json"
EXCEL_FILE = "maigu_sts_compare.xlsx"


# ============================================================
# 数据结构
# ============================================================

@dataclass
class SingleResult:
    endpoint: str       # 端点名（麦谷上游 / STS中转）
    condition: str      # 4K / 16K / 64K / 128K
    model: str
    sample_index: int
    input_tokens_est: int
    max_tokens: int
    output_tokens: int
    ttft: float
    tpot: float
    tps: float
    total_time: float
    task_score: float
    success: bool
    reasoning_chunks: int = 0   # 思考内容 chunk 数（用于诊断 thinking 是否启用）
    error: str = ""


@dataclass
class AggregatedResult:
    endpoint: str
    condition: str
    model: str
    samples: int
    input_tokens: int
    max_tokens: int
    output_tokens_avg: float
    ttft_avg: float
    ttft_min: float
    ttft_max: float
    ttft_std: float
    tpot_avg: float
    tpot_min: float
    tpot_max: float
    tpot_std: float
    tps_avg: float
    tps_min: float
    tps_max: float
    total_time_avg: float
    task_score_avg: float
    success_rate: float
    details: list[SingleResult] = field(default_factory=list)


# ============================================================
# URL 构建
# ============================================================

def _build_chat_url(base_url: str) -> str:
    """根据 base_url 拼接 /v1/chat/completions（兼容已带 /v1 的情况）"""
    base = base_url.rstrip("/")
    if base.endswith("/v1"):
        return f"{base}/chat/completions"
    return f"{base}/v1/chat/completions"


# ============================================================
# 上下文构建
# ============================================================

def build_prompt(target_tokens: int) -> str:
    chars_needed = int(target_tokens * 1.4)
    repeats = chars_needed // len(BASE_TEXT) + 1
    context = (BASE_TEXT * repeats)[:chars_needed]
    return (
        f"以下是关于人工智能的文章内容：\n\n{context}\n\n---\n\n"
        f"{EVALUATION_QUESTION}"
    )


# ============================================================
# API 调用
# ============================================================

def call_api(endpoint: dict, model: str, prompt: str, max_tokens: int) -> dict:
    url = _build_chat_url(endpoint["base_url"])
    headers = {
        "Authorization": f"Bearer {endpoint['api_key']}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": max_tokens,
        "stream": True,
    }

    request_start = time.time()
    first_token_time: Optional[float] = None  # 仅在实际 content 到达时才设置
    last_token_time: Optional[float] = None  # 最后一个内容 chunk 时间
    token_timestamps: list[float] = []
    full_content = ""
    full_reasoning = ""
    reasoning_chunks = 0
    usage_data: dict = {}

    try:
        resp = API_SESSION.post(
            url,
            headers=headers,
            json=payload,
            stream=True,
            timeout=REQUEST_TIMEOUT,
        )
        resp.raise_for_status()
    except Exception as e:
        return {
            "success": False,
            "error": f"请求失败: {e}",
            "ttft": 0.0,
            "tpot": 0.0,
            "tps": 0.0,
            "total_time": time.time() - request_start,
            "output_tokens": 0,
            "content": "",
            "reasoning_chunks": 0,
        }

    for line in resp.iter_lines(decode_unicode=True):
        if not line or not line.startswith("data: "):
            continue
        data_str = line[6:].strip()
        if data_str == "[DONE]":
            break

        try:
            chunk = json.loads(data_str)
        except json.JSONDecodeError:
            continue

        now = time.time()
        choices = chunk.get("choices", [])
        if choices:
            delta = choices[0].get("delta", {})
            content = delta.get("content", "")
            reasoning = delta.get("reasoning_content", "")
            if content:
                # 仅在实际内容 token 到达时设置首 token 时间
                if first_token_time is None:
                    first_token_time = now
                last_token_time = now
                token_timestamps.append(now)
                full_content += content
            if reasoning:
                reasoning_chunks += 1
                full_reasoning += reasoning

        if "usage" in chunk:
            usage_data = chunk.get("usage", {}) or {}

    request_end = time.time()

    if first_token_time is None:
        # 没有可见内容输出
        if reasoning_chunks > 0:
            error_msg = f"仅返回思考内容({reasoning_chunks} chunks)，无可见输出（thinking 未禁用）"
        else:
            error_msg = "未收到任何 token 响应"
        return {
            "success": False,
            "error": error_msg,
            "ttft": 0.0,
            "tpot": 0.0,
            "tps": 0.0,
            "total_time": request_end - request_start,
            "output_tokens": 0,
            "content": "",
            "reasoning_chunks": reasoning_chunks,
        }

    ttft = first_token_time - request_start

    # 计算可见 token 数（排除 thinking）
    completion_tokens = usage_data.get("completion_tokens", 0)
    details = usage_data.get("completion_tokens_details") or {}
    reasoning_tokens = details.get("reasoning_tokens", 0) or usage_data.get("reasoning_tokens", 0)
    visible_tokens = max(completion_tokens - reasoning_tokens, 0)

    # 退化：若 usage 没拆分 reasoning_tokens，但有思考 chunks，则按内容字符估算
    if visible_tokens == 0 and token_timestamps:
        visible_tokens = max(1, int(len(full_content) / 1.4))
    elif visible_tokens == 0 and completion_tokens > 0:
        visible_tokens = completion_tokens

    output_tokens = visible_tokens

    # 用首末内容 chunk 间隔作为生成时间（更准确）
    total_token_time = (last_token_time - first_token_time) if last_token_time else 0
    tpot = (total_token_time / output_tokens * 1000) if output_tokens > 0 else 0
    tps = (output_tokens / total_token_time) if total_token_time > 0 else 0

    return {
        "success": True,
        "error": "",
        "ttft": round(ttft, 3),
        "tpot": round(tpot, 2),
        "tps": round(tps, 2),
        "total_time": round(request_end - request_start, 3),
        "output_tokens": output_tokens,
        "content": full_content,
        "reasoning_chunks": reasoning_chunks,
    }


# ============================================================
# 任务完成质量评估
# ============================================================

def evaluate_task(content: str) -> float:
    if not content:
        return 0.0

    score = 0.0
    # 1. 长度（0-25）
    cl = len(content)
    if cl > 300:
        score += 25
    elif cl > 150:
        score += 20
    elif cl > 80:
        score += 15
    elif cl > 30:
        score += 10
    elif cl > 0:
        score += 5

    # 2. 关键词覆盖（0-30）
    keywords = ["AI", "人工智能", "Transformer", "自注意力", "模型", "学习",
                "数据", "应用", "发展", "挑战"]
    found = sum(1 for kw in keywords if kw in content)
    score += min(found * 3, 30)

    # 3. 结构完整性（0-25）
    stripped = content.rstrip()
    if stripped and stripped[-1] in ("。", "！", "？", ".", "!", "?", "）", ")"):
        score += 12
    if "\n" in content.strip():
        score += 13

    # 4. 任务命中（0-20）— 检查是否包含 3 个子任务要素
    task_hints = 0
    if any(kw in content for kw in ["核心观点", "概括", "总结"]):
        task_hints += 1
    if any(kw in content for kw in ["关键", "信息点", "要点", "1.", "2.", "3."]):
        task_hints += 1
    if "Transformer" in content:
        task_hints += 1
    score += min(task_hints * 7, 20)

    return min(score, 100.0)


# ============================================================
# 单次测试
# ============================================================

def run_single_test(endpoint: dict, model: str, condition: dict,
                     sample_index: int) -> SingleResult:
    prompt = build_prompt(condition["input_tokens"])
    r = call_api(endpoint, model, prompt, condition["max_tokens"])

    task_score = evaluate_task(r["content"]) if r["success"] else 0.0

    return SingleResult(
        endpoint=endpoint["name"],
        condition=condition["name"],
        model=model,
        sample_index=sample_index,
        input_tokens_est=condition["input_tokens"],
        max_tokens=condition["max_tokens"],
        output_tokens=r["output_tokens"],
        ttft=r["ttft"],
        tpot=r["tpot"],
        tps=r["tps"],
        total_time=r["total_time"],
        task_score=task_score,
        success=r["success"],
        reasoning_chunks=r.get("reasoning_chunks", 0),
        error=r["error"],
    )


def _save_progress(progress: list, path: str):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


# ============================================================
# 数据聚合
# ============================================================

def aggregate_results(all_results: list[SingleResult]) -> list[AggregatedResult]:
    groups: dict = defaultdict(list)
    for r in all_results:
        groups[(r.endpoint, r.model, r.condition)].append(r)

    aggregated: list[AggregatedResult] = []
    for (endpoint, model, cond_name), items in sorted(groups.items()):
        success_items = [i for i in items if i.success]
        cond_info = next((c for c in CONDITIONS if c["name"] == cond_name), CONDITIONS[0])

        if success_items:
            output_tokens = [i.output_tokens for i in success_items]
            ttfts = [i.ttft for i in success_items]
            tpots = [i.tpot for i in success_items]
            tpss = [i.tps for i in success_items]
            total_times = [i.total_time for i in success_items]
            task_scores = [i.task_score for i in success_items]
        else:
            output_tokens = ttfts = tpots = tpss = total_times = task_scores = [0]

        aggregated.append(AggregatedResult(
            endpoint=endpoint,
            condition=cond_name,
            model=model,
            samples=len(items),
            input_tokens=cond_info["input_tokens"],
            max_tokens=cond_info["max_tokens"],
            output_tokens_avg=round(statistics.mean(output_tokens), 1),
            ttft_avg=round(statistics.mean(ttfts), 3),
            ttft_min=round(min(ttfts), 3),
            ttft_max=round(max(ttfts), 3),
            ttft_std=round(statistics.stdev(ttfts), 3) if len(ttfts) > 1 else 0,
            tpot_avg=round(statistics.mean(tpots), 2),
            tpot_min=round(min(tpots), 2),
            tpot_max=round(max(tpots), 2),
            tpot_std=round(statistics.stdev(tpots), 2) if len(tpots) > 1 else 0,
            tps_avg=round(statistics.mean(tpss), 2),
            tps_min=round(min(tpss), 2),
            tps_max=round(max(tpss), 2),
            total_time_avg=round(statistics.mean(total_times), 3),
            task_score_avg=round(statistics.mean(task_scores), 1),
            success_rate=round(len(success_items) / len(items) * 100, 1) if items else 0,
            details=items,
        ))

    return aggregated


# ============================================================
# Excel 报告
# ============================================================

def save_excel_report(aggregated: list[AggregatedResult], path: str):
    if not HAS_OPENPYXL:
        print("  [跳过] openpyxl 未安装")
        return

    wb = openpyxl.Workbook()

    # 样式
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    sub_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    sub_font = Font(name="微软雅黑", bold=True, size=10)
    relay_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    upstream_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    data_font = Font(name="微软雅黑", size=10)
    win_font = Font(name="微软雅黑", size=10, bold=True, color="C00000")
    num_fmt = '#,##0.00'
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin

    def style_data(cell, is_num=False):
        cell.font = data_font
        cell.alignment = center
        cell.border = thin
        if is_num:
            cell.number_format = num_fmt

    # ========================================
    # Sheet 1: 对比汇总
    # ========================================
    ws1 = wb.active
    ws1.title = "对比汇总"

    headers = [
        "端点", "模型", "上下文", "样本数",
        "输入Tokens", "输出Tokens(avg)",
        "TTFT(avg/s)", "TTFT(min)", "TTFT(max)", "TTFT(std)",
        "TPOT(avg/ms)", "TPOT(min)", "TPOT(max)", "TPOT(std)",
        "TPS(avg)", "TPS(min)", "TPS(max)",
        "总耗时(avg/s)", "任务评分(avg)", "成功率(%)",
    ]
    for ci, h in enumerate(headers, 1):
        ws1.cell(row=1, column=ci, value=h)
    style_header(ws1, 1, len(headers))

    row = 2
    for a in aggregated:
        values = [
            a.endpoint, a.model, a.condition, a.samples,
            a.input_tokens, a.output_tokens_avg,
            a.ttft_avg, a.ttft_min, a.ttft_max, a.ttft_std,
            a.tpot_avg, a.tpot_min, a.tpot_max, a.tpot_std,
            a.tps_avg, a.tps_min, a.tps_max,
            a.total_time_avg, a.task_score_avg, a.success_rate,
        ]
        for ci, v in enumerate(values, 1):
            cell = ws1.cell(row=row, column=ci, value=v)
            is_num = ci >= 5
            style_data(cell, is_num=is_num)
            # 端点着色
            if ci == 1:
                if a.endpoint == "STS中转":
                    cell.fill = relay_fill
                else:
                    cell.fill = upstream_fill
        row += 1

    ws1.freeze_panes = "A2"
    widths = [12, 12, 10, 8] + [14] * (len(headers) - 4)
    for ci, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ========================================
    # Sheet 2: 中转 vs 上游 并排对比
    # ========================================
    ws2 = wb.create_sheet("中转vs上游")

    compare_headers = [
        "上下文", "模型",
        "TTFT-麦谷(s)", "TTFT-STS(s)", "TTFT差值", "TTFT胜出",
        "TPOT-麦谷(ms)", "TPOT-STS(ms)", "TPOT差值", "TPOT胜出",
        "TPS-麦谷", "TPS-STS", "TPS差值", "TPS胜出",
        "评分-麦谷", "评分-STS",
        "成功率-麦谷", "成功率-STS",
    ]
    for ci, h in enumerate(compare_headers, 1):
        ws2.cell(row=1, column=ci, value=h)
    style_header(ws2, 1, len(compare_headers))

    row = 2
    cond_order = ["4K", "16K", "64K", "128K"]
    for cond_name in cond_order:
        ws2.cell(row=row, column=1, value=f"=== {cond_name} ===")
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(compare_headers))
        for c in range(1, len(compare_headers) + 1):
            cell = ws2.cell(row=row, column=c)
            cell.fill = sub_fill
            cell.font = sub_font
            cell.border = thin
        row += 1

        for model in MODELS:
            up = next((a for a in aggregated
                       if a.endpoint == "麦谷上游" and a.condition == cond_name and a.model == model), None)
            st = next((a for a in aggregated
                       if a.endpoint == "STS中转" and a.condition == cond_name and a.model == model), None)

            up_ttft = up.ttft_avg if up else 0
            st_ttft = st.ttft_avg if st else 0
            ttft_diff = round(st_ttft - up_ttft, 3) if up and st else 0
            ttft_win = "STS快" if (up and st and st_ttft < up_ttft) else ("麦谷快" if (up and st and up_ttft < st_ttft) else "-")

            up_tpot = up.tpot_avg if up else 0
            st_tpot = st.tpot_avg if st else 0
            tpot_diff = round(st_tpot - up_tpot, 2) if up and st else 0
            tpot_win = "STS快" if (up and st and st_tpot < up_tpot) else ("麦谷快" if (up and st and up_tpot < st_tpot) else "-")

            up_tps = up.tps_avg if up else 0
            st_tps = st.tps_avg if st else 0
            tps_diff = round(st_tps - up_tps, 2) if up and st else 0
            tps_win = "STS高" if (up and st and st_tps > up_tps) else ("麦谷高" if (up and st and up_tps > st_tps) else "-")

            values = [
                cond_name, model,
                up_ttft, st_ttft, ttft_diff, ttft_win,
                up_tpot, st_tpot, tpot_diff, tpot_win,
                up_tps, st_tps, tps_diff, tps_win,
                up.task_score_avg if up else 0,
                st.task_score_avg if st else 0,
                up.success_rate if up else 0,
                st.success_rate if st else 0,
            ]
            for ci, v in enumerate(values, 1):
                cell = ws2.cell(row=row, column=ci, value=v)
                is_num = ci in (3, 4, 5, 7, 8, 9, 11, 12, 13, 15, 16, 17, 18)
                style_data(cell, is_num=is_num)
                # 胜出列加粗
                if ci in (6, 10, 14) and v != "-":
                    cell.font = win_font
            row += 1

    ws2.freeze_panes = "A2"
    widths2 = [10, 12, 14, 14, 12, 12, 16, 16, 12, 12, 12, 12, 12, 12, 12, 12, 14, 14]
    for ci, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ========================================
    # Sheet 3: 明细
    # ========================================
    ws3 = wb.create_sheet("明细数据")
    detail_headers = [
        "端点", "模型", "上下文", "样本#",
        "输入Tokens", "输出Tokens",
        "TTFT(s)", "TPOT(ms)", "TPS",
        "总耗时(s)", "任务评分", "思考chunks", "成功", "错误信息",
    ]
    for ci, h in enumerate(detail_headers, 1):
        ws3.cell(row=1, column=ci, value=h)
    style_header(ws3, 1, len(detail_headers))

    row = 2
    for a in aggregated:
        for d in a.details:
            values = [
                d.endpoint, d.model, d.condition, d.sample_index + 1,
                d.input_tokens_est, d.output_tokens,
                d.ttft, d.tpot, d.tps,
                d.total_time, d.task_score, d.reasoning_chunks,
                "✓" if d.success else "✗", d.error,
            ]
            for ci, v in enumerate(values, 1):
                cell = ws3.cell(row=row, column=ci, value=v)
                is_num = ci in (5, 6, 7, 8, 9, 10, 11, 12)
                style_data(cell, is_num=is_num)
                if ci == 1:
                    if d.endpoint == "STS中转":
                        cell.fill = relay_fill
                    else:
                        cell.fill = upstream_fill
            row += 1

    ws3.freeze_panes = "A2"
    for ci, w in enumerate([12, 12, 10, 8, 14, 14, 10, 10, 8, 12, 10, 12, 8, 40], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    wb.save(path)
    print(f"  Excel 报告已保存: {path}")


# ============================================================
# JSON 报告
# ============================================================

def save_json_report(aggregated: list[AggregatedResult], path: str):
    data = {
        "generated_at": datetime.now().isoformat(),
        "endpoints": [e["name"] for e in ENDPOINTS],
        "models": MODELS,
        "conditions": [c["name"] for c in CONDITIONS],
        "results": [],
    }
    for a in aggregated:
        d = asdict(a)
        d.pop("details", None)
        data["results"].append(d)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  JSON 报告已保存: {path}")


# ============================================================
# 终端摘要
# ============================================================

def print_summary(aggregated: list[AggregatedResult]):
    print("\n" + "=" * 90)
    print("📊 麦谷上游 vs STS中转 对比报告")
    print("=" * 90)

    cond_order = ["4K", "16K", "64K", "128K"]
    for cond_name in cond_order:
        print(f"\n{'─' * 88}")
        print(f"  📌 上下文: {cond_name}")
        print(f"{'─' * 88}")
        print(f"  {'端点':<10} {'模型':<10} {'TTFT(s)':<10} {'TPOT(ms)':<10} "
              f"{'TPS':<8} {'任务评分':<10} {'成功率':<8}")
        print(f"  {'─' * 78}")
        items = [a for a in aggregated if a.condition == cond_name]
        for a in items:
            print(f"  {a.endpoint:<10} {a.model:<10} "
                  f"{a.ttft_avg:<10.3f} {a.tpot_avg:<10.2f} "
                  f"{a.tps_avg:<8.2f} {a.task_score_avg:<10.1f} {a.success_rate:<8.1f}")


# ============================================================
# 主流程
# ============================================================

def main():
    import sys
    quick_mode = "--quick" in sys.argv
    resume_mode = "--resume" in sys.argv

    # 续跑加载
    existing: list[dict] = []
    if resume_mode and os.path.isfile(REPORT_FILE):
        try:
            with open(REPORT_FILE, encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list):
                existing = data
            elif isinstance(data, dict):
                existing = data.get("details", data.get("results", []))
            print(f"[信息] 已加载 {len(existing)} 条现有结果（断点续跑）")
        except Exception:
            pass

    completed = set()
    for r in existing:
        completed.add((r.get("endpoint"), r.get("model"),
                       r.get("condition"), r.get("sample_index")))

    print("=" * 70)
    print("🔬 麦谷上游 vs STS中转 GLM 对比测试")
    print("=" * 70)
    print(f"  端点: {[e['name'] for e in ENDPOINTS]}")
    print(f"  模型: {MODELS}")
    print(f"  上下文: {[c['name'] for c in CONDITIONS]}")
    mode_str = "快速" if quick_mode else ("续跑" if resume_mode else "完整")
    print(f"  模式: {mode_str}")
    print()

    all_results: list[SingleResult] = []
    for r in existing:
        all_results.append(SingleResult(
            endpoint=r["endpoint"], condition=r["condition"], model=r["model"],
            sample_index=r["sample_index"], input_tokens_est=r["input_tokens_est"],
            max_tokens=r["max_tokens"], output_tokens=r["output_tokens"],
            ttft=r["ttft"], tpot=r["tpot"], tps=r["tps"],
            total_time=r["total_time"], task_score=r["task_score"],
            success=r["success"], reasoning_chunks=r.get("reasoning_chunks", 0),
            error=r.get("error", ""),
        ))

    for endpoint in ENDPOINTS:
        for model in MODELS:
            for condition in CONDITIONS:
                cond_name = condition["name"]
                samples = 1 if quick_mode else condition["samples"]

                print(f"▶  [{endpoint['name']}] [{model}] [{cond_name}] ...")

                for sample_idx in range(samples):
                    key = (endpoint["name"], model, cond_name, sample_idx)
                    if key in completed:
                        continue

                    try:
                        r = run_single_test(endpoint, model, condition, sample_idx)
                        all_results.append(r)
                        completed.add(key)
                        _save_progress([asdict(x) for x in all_results], REPORT_FILE)

                        status = "✓" if r.success else "✗"
                        reasoning_tag = f" 思考={r.reasoning_chunks}chunks" if r.reasoning_chunks > 0 else ""
                        print(f"    [{cond_name}] 样本#{sample_idx + 1} {status} "
                              f"TTFT={r.ttft:.3f}s TPOT={r.tpot:.2f}ms "
                              f"TPS={r.tps:.2f} 评分={r.task_score:.1f}{reasoning_tag}"
                              + (f" [{r.error}]" if not r.success else ""))
                    except Exception as e:
                        print(f"    [{cond_name}] 样本#{sample_idx + 1} ✗ 异常: {e}")
                        traceback.print_exc()

    print("\n[信息] 正在聚合数据...")
    aggregated = aggregate_results(all_results)

    print("\n[信息] 正在生成报告...")
    save_json_report(aggregated, REPORT_FILE)
    if HAS_OPENPYXL:
        save_excel_report(aggregated, EXCEL_FILE)

    print_summary(aggregated)

    print("\n" + "=" * 70)
    print("✅ 测试完成！")
    print(f"  JSON 报告: {REPORT_FILE}")
    if HAS_OPENPYXL:
        print(f"  Excel 报告: {EXCEL_FILE}")
    print("=" * 70)


if __name__ == "__main__":
    main()
