"""
TokenEasy API Claude 模型基准测试脚本
=========================================
测试 8 个 Claude 模型在 4K/16K/64K/128K 上下文下的性能指标：
  - TTFT (Time to First Token)       首 token 延迟
  - TPOT (Time per Output Token)      每 token 输出时间
  - TPS  (Tokens Per Second)          每秒输出 token 数
  - 任务完成质量评分
  - 生成 Excel 报告

用法:
  python tokeneasy_benchmark.py          # 完整测试
  python tokeneasy_benchmark.py --quick  # 快速模式（每组合跑1次，14K只用1次）
  python tokeneasy_benchmark.py --resume # 断点续跑
"""

from __future__ import annotations

import json
import os
import statistics
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field, asdict
from datetime import datetime
from typing import Optional

import requests

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[警告] openpyxl 未安装，无法生成 Excel。请执行: pip install openpyxl")


# ============================================================
# 配置
# ============================================================

# 从 .env 加载 TokenEasy 配置
def _load_env() -> tuple[str, str]:
    base_url = "https://api.tokeneasy.ai"
    api_key = ""
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    if os.path.isfile(env_path):
        with open(env_path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line.startswith("TOKENASYNC_BASE_URL="):
                    base_url = line.split("=", 1)[1].strip().strip('"').strip("'")
                elif line.startswith("TOKENASYNC_API_KEY="):
                    api_key = line.split("=", 1)[1].strip().strip('"').strip("'")
    return base_url, api_key


TOKENASYNC_BASE_URL, TOKENASYNC_API_KEY = _load_env()
if not TOKENASYNC_API_KEY:
    print("[错误] 请在 .env 中设置 TOKENASYNC_API_KEY")
    exit(1)

# 测试的模型列表
MODELS: list[str] = [
    "claude-fable-5",
    "claude-sonnet-4-5",
    "claude-sonnet-5",
    "claude-opus-4-8",
    "claude-haiku-4-5",
    "claude-sonnet-4-6",
    "claude-opus-4-7",
    "claude-opus-4-6",
]

# 上下文测试条件
CONDITIONS: list[dict] = [
    {"name": "4K",   "input_tokens": 3072,  "max_tokens": 1024, "samples": 3},
    {"name": "16K",  "input_tokens": 14000, "max_tokens": 2048, "samples": 3},
    {"name": "64K",  "input_tokens": 62000, "max_tokens": 2048, "samples": 2},
    {"name": "128K", "input_tokens": 126000,"max_tokens": 2048, "samples": 2},
]

# 并发数
CONCURRENT_WORKERS = 2

# 请求超时（秒）
REQUEST_TIMEOUT = 300

# 任务评估 prompt（用于检测模型完成任务的能力）
EVALUATION_QUESTIONS: list[dict] = [
    {
        "instruction": "请用一句话概括下面这段文字的核心观点。",
        "max_score": 25,
    },
    {
        "instruction": "请从这段文字中列出3个关键信息点。",
        "max_score": 25,
    },
    {
        "instruction": "请对文中提到的 AI 技术趋势做一个简要分类总结。",
        "max_score": 25,
    },
    {
        "instruction": "请根据文章内容，提出2个值得深入思考的问题。",
        "max_score": 25,
    },
]

# 背景语料（用于填充上下文；1 token ≈ 1.4 中文字符）
BASE_TEXT = (
    "人工智能（Artificial Intelligence，简称AI）是计算机科学的一个重要分支，"
    "它致力于研究、开发用于模拟、延伸和扩展人类智能的理论、方法、技术及应用系统。"
    "自1956年达特茅斯会议正式提出人工智能概念以来，该领域经历了多次起伏，"
    "从早期的符号主义到专家系统，再到统计机器学习和深度学习的崛起。"
    "近年来，随着深度学习技术的突破、大数据基础设施的发展以及GPU/TPU等计算能力的提升，"
    "人工智能在图像识别、自然语言处理、语音识别、自动驾驶、医疗诊断等领域取得了显著进展。"
    "大型语言模型（LLM）如GPT系列、GLM系列、Claude系列等的出现，更是将人工智能推向了一个新的高度，"
    "这些模型通过对海量文本数据的学习，展现出了强大的语言理解和生成能力。"
    "在自然语言处理领域，预训练-微调范式已成为主流，BERT、GPT、T5等模型架构层出不穷。"
    "Transformer架构的自注意力机制使得模型能够捕捉长距离依赖关系，这是传统RNN/LSTM难以做到的。"
    "强化学习与人类反馈（RLHF）技术的引入，使得大语言模型能够更好地对齐人类意图和价值观。"
    "然而，人工智能的发展也面临着诸多挑战，包括数据隐私保护、算法公平性与偏见、能源消耗、"
    "以及如何确保AI系统的安全性、可控性和可解释性等问题。"
    "在未来，人工智能将继续深刻影响各行各业，从医疗健康到教育培训，"
    "从金融科技到智能制造，从智慧城市到环境保护，AI的应用前景广阔无垠。"
    "同时，我们也需要关注AI伦理和治理，确保技术的发展真正造福人类社会。"
    "量子计算与人工智能的结合可能带来革命性突破，量子机器学习有望解决经典计算难以处理的问题。"
    "边缘计算与联邦学习则让AI能力延伸到更多终端设备，同时保护用户数据隐私。"
    "多模态学习使得AI能够同时处理文本、图像、音频、视频等多种信息形式，"
    "这为实现更接近人类的感知和理解能力奠定了基础。"
    "AI Agent技术的发展使得大模型能够自主规划、调用工具、执行多步骤任务，"
    "这标志着人工智能从被动应答向主动执行的重要转变。"
    "在实际应用中，AI辅助编程工具如GitHub Copilot、Cursor等极大地提高了开发者的生产效率。"
    "AI在科学研究中的应用也日益广泛，从蛋白质结构预测到新材料发现，从气候模拟到药物研发。"
    "计算机视觉领域，目标检测、语义分割、图像生成（如Stable Diffusion、DALL-E）等技术日趋成熟。"
    "语音技术方面，语音识别、语音合成、声纹识别等已广泛应用于智能音箱、客服系统等场景。"
    "推荐系统借助深度学习，在电商、短视频、社交媒体等平台实现了精准的个性化推荐。"
    "知识图谱与大型语言模型的结合，有望解决大模型的幻觉问题和知识更新问题。"
    "检索增强生成（RAG）技术通过引入外部知识库，显著提升了模型回答的准确性和时效性。"
    "提示工程（Prompt Engineering）和上下文学习（In-Context Learning）成为高效使用大模型的关键技能。"
    "模型压缩技术如量化、剪枝、蒸馏等，使得大模型能够在资源受限的设备上高效运行。"
)

REPORT_FILE = "tokeneasy_report.json"
EXCEL_FILE = "tokeneasy_report.xlsx"


# ============================================================
# 数据结构
# ============================================================

@dataclass
class SingleResult:
    """单次请求的测试结果"""
    condition: str            # 如 "4K", "16K"
    model: str                # 模型名
    sample_index: int         # 第几次重复
    input_tokens_est: int     # 估计输入 token 数
    max_tokens: int           # 最大输出 token 数
    output_tokens: int        # 实际输出 token 数
    ttft: float               # 首 token 延迟（秒）
    tpot: float               # 每 token 输出时间（ms）
    tps: float                # 每秒输出 token 数
    total_time: float         # 总耗时（秒）
    success: bool             # 是否成功
    question_idx: int = 0     # 问题序号
    task_score_avg: float = 0.0  # 任务评分
    task_scores: list[int] | None = None  # 各任务评分
    error: str = ""           # 错误信息


@dataclass
class AggregatedResult:
    """同一个模型×同一种条件聚合后的结果"""
    condition: str
    model: str
    samples: int
    input_tokens: int
    max_tokens: int
    output_tokens_avg: float
    output_tokens_min: float
    output_tokens_max: float
    ttft_avg: float
    ttft_min: float
    ttft_max: float
    ttft_std: float
    tpot_avg: float
    tpot_min: float
    tpot_max: float
    tpot_std: float
    tps_avg: float
    tps_min: float
    tps_max: float
    total_time_avg: float
    task_score_avg: float
    task_score_min: float
    task_score_max: float
    success_rate: float
    details: list[SingleResult] = field(default_factory=list)


# ============================================================
# 填充上下文构建
# ============================================================

def build_prompt(target_tokens: int, question: str) -> str:
    """构建指定 token 数的背景 + 问题"""
    # 估算：1 token ≈ 1.4 中文字符
    chars_needed = int(target_tokens * 1.4)
    repeats = chars_needed // len(BASE_TEXT) + 1
    context = (BASE_TEXT * repeats)[:chars_needed]
    return f"以下是关于人工智能的文章内容：\n\n{context}\n\n---\n\n{question}"


# ============================================================
# API 调用与计时
# ============================================================

def call_api(model: str, prompt: str, max_tokens: int) -> dict:
    """
    调用 TokenEasy API，返回计时数据。
    返回字典包含 tokens 统计和计时信息。
    """
    headers = {
        "Authorization": f"Bearer {TOKENASYNC_API_KEY}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": max_tokens,
        "stream": True,
    }

    request_start = time.time()
    first_token_time: Optional[float] = None
    token_timestamps: list[float] = []
    full_content = ""
    response_data = {}

    try:
        resp = requests.post(
            f"{TOKENASYNC_BASE_URL}/v1/chat/completions",
            headers=headers,
            json=payload,
            stream=True,
            timeout=REQUEST_TIMEOUT,
        )
        resp.raise_for_status()
    except Exception as e:
        return {
            "success": False,
            "error": f"请求失败: {e}",
            "ttft": 0,
            "tpot": 0,
            "tps": 0,
            "total_time": time.time() - request_start,
            "output_tokens": 0,
            "content": "",
        }

    # 解析 SSE 流
    for line in resp.iter_lines(decode_unicode=True):
        if not line or not line.startswith("data: "):
            continue
        data_str = line[6:].strip()
        if data_str == "[DONE]":
            break

        try:
            chunk = json.loads(data_str)
        except json.JSONDecodeError:
            continue

        now = time.time()

        # 首 token 计时
        if first_token_time is None:
            first_token_time = now

        choices = chunk.get("choices", [])
        if choices:
            delta = choices[0].get("delta", {})
            content = delta.get("content", "")
            if content:
                token_timestamps.append(now)
                full_content += content

        # 收集 usage 信息（最后一条 chunk 可能携带，choices为空数组）
        if "usage" in chunk:
            response_data = chunk.get("usage", {}) or {}

    request_end = time.time()

    if first_token_time is None:
        return {
            "success": False,
            "error": "未收到任何 token 响应",
            "ttft": 0,
            "tpot": 0,
            "tps": 0,
            "total_time": request_end - request_start,
            "output_tokens": 0,
            "content": "",
        }

    ttft = first_token_time - request_start

    # 输出 token 数
    output_tokens = response_data.get("completion_tokens", 0)
    if output_tokens <= 0 and token_timestamps:
        # 估算：中文字符数 / 1.4
        output_tokens = max(1, int(len(full_content) / 1.4))

    total_token_time = request_end - first_token_time if token_timestamps else 0
    tpot = (total_token_time / output_tokens * 1000) if output_tokens > 0 else 0
    tps = (output_tokens / total_token_time) if total_token_time > 0 else 0

    return {
        "success": True,
        "error": "",
        "ttft": round(ttft, 3),
        "tpot": round(tpot, 2),
        "tps": round(tps, 2),
        "total_time": round(request_end - request_start, 3),
        "output_tokens": output_tokens,
        "content": full_content,
    }


# ============================================================
# 任务完成质量评估
# ============================================================

def evaluate_task(content: str, expected: str) -> int:
    """
    评估模型回答质量。
    基于回答长度、相关性和完整性做简单打分。
    """
    if not content:
        return 0

    score = 0

    # 1. 回答非空（0-20）
    content_len = len(content)
    if content_len > 200:
        score += 20
    elif content_len > 100:
        score += 15
    elif content_len > 50:
        score += 10
    elif content_len > 10:
        score += 5

    # 2. 包含关键要素（0-40）
    keywords = ["AI", "人工智能", "技术", "模型", "数据", "学习", "应用", "发展", "未来", "挑战"]
    found = sum(1 for kw in keywords if kw in content)
    score += min(found * 4, 40)

    # 3. 结构完整（0-20）
    has_punctuation_end = content.rstrip()[-1] in ("。", "！", "？", ".", "!", "?", "）", ")") if content.rstrip() else False
    has_newline = "\n" in content.strip()
    if has_punctuation_end:
        score += 10
    if has_newline:
        score += 10

    # 4. 与问题相关（0-20）
    expected_keywords = expected.replace("请", "").replace("?", " ").replace("？", " ")
    expected_words = set(expected_keywords.split())
    relevant = sum(1 for w in expected_words if len(w) > 1 and w in content)
    score += min(relevant * 5, 20)

    return min(score, 100)


# ============================================================
# 单次测试
# ============================================================

def run_single_test(
    model: str,
    condition: dict,
    sample_index: int,
    question: str,
    question_idx: int,
) -> dict:
    """跑一次请求并返回结果"""
    target_tokens = condition["input_tokens"]
    max_tokens = condition["max_tokens"]
    prompt = build_prompt(target_tokens, question)

    result = call_api(model, prompt, max_tokens)

    task_score = 0
    if result["success"]:
        task_score = evaluate_task(result["content"], question)

    return {
        "condition": condition["name"],
        "model": model,
        "sample_index": sample_index,
        "question_idx": question_idx,
        "input_tokens_est": target_tokens,
        "max_tokens": max_tokens,
        "output_tokens": result["output_tokens"],
        "ttft": result["ttft"],
        "tpot": result["tpot"],
        "tps": result["tps"],
        "total_time": result["total_time"],
        "task_score": task_score,
        "success": result["success"],
        "error": result["error"],
    }


def run_model_condition(
    model: str,
    condition: dict,
    progress: list,
    results_file: str,
) -> list[dict]:
    """对一个模型×条件组合运行所有评估问题×重复次数"""
    cond_name = condition["name"]
    samples = condition["samples"]
    all_results: list[dict] = []

    for sample_idx in range(samples):
        for qi, q in enumerate(EVALUATION_QUESTIONS):
            r = run_single_test(
                model=model,
                condition=condition,
                sample_index=sample_idx,
                question=q["instruction"],
                question_idx=qi,
            )
            all_results.append(r)

            # 保存进度
            progress.append(r)
            _save_progress(progress, results_file)

    return all_results


def _save_progress(progress: list, path: str):
    """增量保存到 JSON"""
    with open(path, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


# ============================================================
# 数据聚合
# ============================================================

def aggregate_results(all_results: list[dict]) -> list[AggregatedResult]:
    """将原始结果按 (model, condition) 分组聚合"""
    from collections import defaultdict

    groups: defaultdict = defaultdict(list)
    for r in all_results:
        key = (r["model"], r["condition"])
        groups[key].append(r)

    aggregated: list[AggregatedResult] = []
    for (model, cond_name), items in sorted(groups.items()):
        success_items = [i for i in items if i["success"]]

        # 条件信息
        cond_info = next((c for c in CONDITIONS if c["name"] == cond_name), CONDITIONS[0])

        if success_items:
            output_tokens = [i["output_tokens"] for i in success_items]
            ttfts = [i["ttft"] for i in success_items]
            tpots = [i["tpot"] for i in success_items]
            tpss = [i["tps"] for i in success_items]
            total_times = [i["total_time"] for i in success_items]
            task_scores = [i["task_score"] for i in success_items]
        else:
            output_tokens = ttfts = tpots = tpss = total_times = task_scores = [0]

        agg = AggregatedResult(
            condition=cond_name,
            model=model,
            samples=len(items),
            input_tokens=cond_info["input_tokens"],
            max_tokens=cond_info["max_tokens"],
            output_tokens_avg=round(statistics.mean(output_tokens), 1),
            output_tokens_min=round(min(output_tokens), 1),
            output_tokens_max=round(max(output_tokens), 1),
            ttft_avg=round(statistics.mean(ttfts), 3),
            ttft_min=round(min(ttfts), 3),
            ttft_max=round(max(ttfts), 3),
            ttft_std=round(statistics.stdev(ttfts), 3) if len(ttfts) > 1 else 0,
            tpot_avg=round(statistics.mean(tpots), 2),
            tpot_min=round(min(tpots), 2),
            tpot_max=round(max(tpots), 2),
            tpot_std=round(statistics.stdev(tpots), 2) if len(tpots) > 1 else 0,
            tps_avg=round(statistics.mean(tpss), 2),
            tps_min=round(min(tpss), 2),
            tps_max=round(max(tpss), 2),
            total_time_avg=round(statistics.mean(total_times), 3),
            task_score_avg=round(statistics.mean(task_scores), 1),
            task_score_min=round(min(task_scores), 1),
            task_score_max=round(max(task_scores), 1),
            success_rate=round(len(success_items) / len(items) * 100, 1) if items else 0,
            details=[SingleResult(
                condition=i["condition"],
                model=i["model"],
                sample_index=i["sample_index"],
                question_idx=i.get("question_idx", 0),
                input_tokens_est=i["input_tokens_est"],
                max_tokens=i["max_tokens"],
                output_tokens=i["output_tokens"],
                ttft=i["ttft"],
                tpot=i["tpot"],
                tps=i["tps"],
                total_time=i["total_time"],
                task_scores=[i["task_score"]],
                task_score_avg=float(i["task_score"]),
                success=i["success"],
                error=i.get("error", ""),
            ) for i in items],
        )
        aggregated.append(agg)

    return aggregated


# ============================================================
# 报告输出
# ============================================================

def save_json_report(aggregated: list[AggregatedResult], path: str):
    """保存 JSON 报告"""
    data = {
        "generated_at": datetime.now().isoformat(),
        "total_models": len(MODELS),
        "total_conditions": len(CONDITIONS),
        "results": [asdict(a) for a in aggregated],
    }
    # 去掉 details 中的大量数据
    for r in data["results"]:
        r.pop("details", None)

    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  JSON 报告已保存: {path}")


def save_excel_report(aggregated: list[AggregatedResult], path: str):
    """生成格式化的 Excel 报告"""
    if not HAS_OPENPYXL:
        print("  [跳过] openpyxl 未安装")
        return

    wb = openpyxl.Workbook()

    # ---------- 样式 ----------
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    subheader_font = Font(name="微软雅黑", bold=True, size=10)
    data_font = Font(name="微软雅黑", size=10)
    num_fmt = '#,##0.00'
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

    def style_data_cell(cell, is_num=False):
        cell.font = data_font
        cell.alignment = center_align
        cell.border = thin_border
        if is_num:
            cell.number_format = num_fmt

    # ========================================
    # Sheet 1: 汇总表（模型 × 条件）
    # ========================================
    ws1 = wb.active
    ws1.title = "汇总"

    headers = [
        "模型", "上下文", "样本数",
        "输入 Tokens", "输出 Tokens(avg)",
        "TTFT(avg/s)", "TTFT(min)", "TTFT(max)", "TTFT(std)",
        "TPOT(avg/ms)", "TPOT(min)", "TPOT(max)", "TPOT(std)",
        "TPS(avg)", "TPS(min)", "TPS(max)",
        "总耗时(avg/s)",
        "任务评分(avg)", "任务评分(min)", "任务评分(max)",
        "成功率(%)",
    ]

    for ci, h in enumerate(headers, 1):
        ws1.cell(row=1, column=ci, value=h)
    style_header(ws1, 1, len(headers))

    row = 2
    for a in aggregated:
        values = [
            a.model, a.condition, a.samples,
            a.input_tokens, a.output_tokens_avg,
            a.ttft_avg, a.ttft_min, a.ttft_max, a.ttft_std,
            a.tpot_avg, a.tpot_min, a.tpot_max, a.tpot_std,
            a.tps_avg, a.tps_min, a.tps_max,
            a.total_time_avg,
            a.task_score_avg, a.task_score_min, a.task_score_max,
            a.success_rate,
        ]
        for ci, v in enumerate(values, 1):
            cell = ws1.cell(row=row, column=ci, value=v)
            is_num = ci >= 4
            style_data_cell(cell, is_num=is_num)
        row += 1

    # 冻结首行
    ws1.freeze_panes = "A2"

    # 列宽
    col_widths = [22] + [14] * (len(headers) - 1)
    for ci, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ========================================
    # Sheet 2: 模型对比（每个条件一个块）
    # ========================================
    ws2 = wb.create_sheet("模型对比")

    # 表头
    compare_headers = [
        "上下文", "模型", "TTFT(s)", "TPOT(ms)", "TPS",
        "输出Tokens", "总耗时(s)", "任务评分", "成功率(%)",
    ]
    for ci, h in enumerate(compare_headers, 1):
        ws2.cell(row=1, column=ci, value=h)
    style_header(ws2, 1, len(compare_headers))

    row = 2
    cond_order = ["4K", "16K", "64K", "128K"]
    for cond_name in cond_order:
        items_in_cond = [a for a in aggregated if a.condition == cond_name]
        if not items_in_cond:
            continue

        # 条件标题行
        ws2.cell(row=row, column=1, value=f"=== {cond_name} ===")
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(compare_headers))
        for c in range(1, len(compare_headers) + 1):
            cell = ws2.cell(row=row, column=c)
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.border = thin_border
        row += 1

        for a in items_in_cond:
            values = [
                a.condition, a.model,
                a.ttft_avg, a.tpot_avg, a.tps_avg,
                a.output_tokens_avg, a.total_time_avg,
                a.task_score_avg, a.success_rate,
            ]
            for ci, v in enumerate(values, 1):
                cell = ws2.cell(row=row, column=ci, value=v)
                is_num = ci >= 3
                style_data_cell(cell, is_num=is_num)
            row += 1

    ws2.freeze_panes = "A2"
    for ci, w in enumerate([12, 22, 12, 12, 10, 14, 12, 12, 10], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ========================================
    # Sheet 3: 单次明细
    # ========================================
    ws3 = wb.create_sheet("明细数据")

    detail_headers = [
        "模型", "上下文", "样本#", "问题#",
        "输入Tokens", "输出Tokens",
        "TTFT(s)", "TPOT(ms)", "TPS",
        "总耗时(s)", "任务评分", "成功", "错误信息",
    ]
    for ci, h in enumerate(detail_headers, 1):
        ws3.cell(row=1, column=ci, value=h)
    style_header(ws3, 1, len(detail_headers))

    row = 2
    for a in aggregated:
        for d in a.details:
            values = [
                d.model, d.condition, d.sample_index + 1, d.question_idx + 1,
                d.input_tokens_est, d.output_tokens,
                d.ttft, d.tpot, d.tps,
                d.total_time, d.task_score_avg, "✓" if d.success else "✗", d.error,
            ]
            for ci, v in enumerate(values, 1):
                cell = ws3.cell(row=row, column=ci, value=v)
                is_num = ci in (5, 6, 7, 8, 9, 10, 11)
                style_data_cell(cell, is_num=is_num)
            row += 1

    ws3.freeze_panes = "A2"
    for ci, w in enumerate([22, 10, 8, 8, 14, 14, 10, 10, 8, 12, 10, 8, 40], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    wb.save(path)
    print(f"  Excel 报告已保存: {path}")


# ============================================================
# 主流程
# ============================================================

def print_summary(aggregated: list[AggregatedResult]):
    """打印终端摘要"""
    print("\n" + "=" * 80)
    print("📊 TokenEasy Claude 模型基准测试报告")
    print("=" * 80)

    cond_order = ["4K", "16K", "64K", "128K"]
    for cond_name in cond_order:
        print(f"\n{'─' * 60}")
        print(f"  📌 上下文: {cond_name}")
        print(f"{'─' * 60}")
        print(f"  {'模型':<22} {'TTFT(s)':<10} {'TPOT(ms)':<10} {'TPS':<8} {'任务评分':<10} {'成功率':<8}")
        print(f"  {'─' * 68}")

        items = [a for a in aggregated if a.condition == cond_name]
        for a in sorted(items, key=lambda x: x.tpot_avg):
            print(
                f"  {a.model:<22} "
                f"{a.ttft_avg:<10.3f} "
                f"{a.tpot_avg:<10.2f} "
                f"{a.tps_avg:<8.2f} "
                f"{a.task_score_avg:<10.1f} "
                f"{a.success_rate:<8.1f}"
            )


def main():
    import sys

    # 解析参数
    quick_mode = "--quick" in sys.argv
    resume_mode = "--resume" in sys.argv

    report_file = REPORT_FILE

    # 加载已有进度（断点续跑）
    existing_results: list[dict] = []
    if resume_mode and os.path.isfile(report_file):
        try:
            with open(report_file, encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list):
                existing_results = data
            elif isinstance(data, dict):
                existing_results = data.get("results", [])
            print(f"[信息] 已加载 {len(existing_results)} 条现有结果（断点续跑）")
        except Exception:
            pass

    # 构建已完成集合
    completed = set()
    for r in existing_results:
        completed.add((r["model"], r["condition"], r["sample_index"], r.get("question_idx", 0)))

    print("=" * 60)
    print("🔬 TokenEasy API Claude 模型基准测试")
    print("=" * 60)
    print(f"  API: {TOKENASYNC_BASE_URL}")
    print(f"  模型数: {len(MODELS)}")
    print(f"  上下文条件: {[c['name'] for c in CONDITIONS]}")
    mode_str = "快速" if quick_mode else ("续跑" if resume_mode else "完整")
    print(f"  模式: {mode_str}")
    print()

    all_results: list[dict] = list(existing_results)

    for model in MODELS:
        for condition in CONDITIONS:
            cond_name = condition["name"]
            samples = 1 if quick_mode else condition["samples"]

            print(f"▶  [{model}] [{cond_name}] ...")

            for sample_idx in range(samples):
                for qi, q in enumerate(EVALUATION_QUESTIONS):
                    key = (model, cond_name, sample_idx, qi)
                    if key in completed:
                        continue

                    try:
                        r = run_single_test(
                            model=model,
                            condition=condition,
                            sample_index=sample_idx,
                            question=q["instruction"],
                            question_idx=qi,
                        )
                        all_results.append(r)
                        completed.add(key)
                        _save_progress(all_results, report_file)

                        status = "✓" if r["success"] else "✗"
                        print(f"    [{cond_name}] 样本#{sample_idx + 1} 问题#{qi + 1} "
                              f"{status} TTFT={r['ttft']:.3f}s TPOT={r['tpot']:.2f}ms "
                              f"评分={r['task_score']}")
                    except Exception as e:
                        print(f"    [{cond_name}] 样本#{sample_idx + 1} 问题#{qi + 1} "
                              f"✗ 异常: {e}")
                        traceback.print_exc()

    # 聚合
    print("\n[信息] 正在聚合数据...")
    aggregated = aggregate_results(all_results)

    # 输出
    print("\n[信息] 正在生成报告...")
    save_json_report(aggregated, report_file)

    if HAS_OPENPYXL:
        save_excel_report(aggregated, EXCEL_FILE)

    print_summary(aggregated)

    print("\n" + "=" * 60)
    print("✅ 测试完成！")
    print(f"  JSON 报告: {REPORT_FILE}")
    if HAS_OPENPYXL:
        print(f"  Excel 报告: {EXCEL_FILE}")
    print("=" * 60)


if __name__ == "__main__":
    main()
